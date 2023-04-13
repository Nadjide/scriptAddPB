import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import DateEntry
from openpyxl import load_workbook
from datetime import date as dt
import os
import re

def ajouter_donnees():
    date = date_entry.get()
    probleme = probleme_entry.get("1.0", tk.END).strip()
    numero = numero_entry.get()

    date_pattern = r'^\d{2}/\d{2}/\d{4}$'
    if not re.match(date_pattern, date):
        messagebox.showerror("Erreur", "La date doit être au format jj/mm/aaaa.")
        return

    numero_pattern = r'^\d{2}$'
    if not re.match(numero_pattern, numero):
        messagebox.showerror("Erreur", "Le numéro de poste doit comporter 2 chiffres.")
        return

    chemin_fichier = r'C:\Users\LPT_Nadjide\OneDrive\Documents\PyTest'
    nom_fichier = 'Suivi_PB_Plateau.xlsx'
    fichier_complet = os.path.join(chemin_fichier, nom_fichier)

    wb = load_workbook(fichier_complet)
    ws = wb.active

    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=date)
    ws.cell(row=row, column=2, value=probleme)
    ws.cell(row=row, column=3, value=numero)

    wb.save(fichier_complet)
    os.startfile(fichier_complet)

    annuler_donnees()

    messagebox.showinfo("Succès", "Le problème a été ajouté avec succès !")

def annuler_donnees():
    date_entry.set_date(dt.today())
    probleme_entry.delete("1.0", tk.END)
    numero_entry.delete(0, tk.END)

fenetre = tk.Tk()
fenetre.title("Ajouter un problème")

frame = ttk.Frame(fenetre, padding=10)
frame.grid(row=0, column=0, padx=10, pady=10, sticky=tk.W+tk.E)

date_label = ttk.Label(frame, text="Date (jj/mm/aaaa) :")
date_label.grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
date_entry = DateEntry(frame, width=18, background='blue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
date_entry.grid(row=0, column=1)

probleme_label = ttk.Label(frame, text="Nature du problème :")
probleme_label.grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(10, 0))
probleme_entry = tk.Text(frame, width=20, height=4, wrap=tk.WORD)
probleme_entry.grid(row=1, column=1, pady=(10, 0))

numero_label = ttk.Label(frame, text="Numéro de Poste :")
numero_label.grid(row=2, column=0, sticky=tk.W, padx=(0, 10), pady=(10, 0))
numero_entry = ttk.Entry(frame, width=20)
numero_entry.grid(row=2, column=1, pady=(10, 0))

submit_button = ttk.Button(frame, text="Ajouter", command=ajouter_donnees)
submit_button.grid(row=3, column=1, sticky=tk.E, pady=(20, 0))

cancel_button = ttk.Button(frame, text="Annuler", command=annuler_donnees)
cancel_button.grid(row=3, column=0, sticky=tk.W, pady=(20, 0))

fenetre.columnconfigure(0, weight=1)
fenetre.rowconfigure(0, weight=1)
frame.columnconfigure(1, weight=1)

fenetre.mainloop()
